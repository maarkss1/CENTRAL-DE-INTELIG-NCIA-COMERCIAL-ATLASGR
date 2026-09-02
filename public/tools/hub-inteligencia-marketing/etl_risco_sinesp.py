#!/usr/bin/env python3
"""
Atlas GR Market Intelligence - ETL Sinesp VDE -> risco municipal/UF

Objetivo
--------
Normalizar a Base de Dados do Validador de Dados Estatísticos (Sinesp VDE)
para um CSV compacto que o Atlas Market Intelligence consegue importar.

Fonte oficial sugerida
----------------------
MJSP / Sinesp - Base de Dados VDE (2015-2026):
https://dados.mj.gov.br/dataset/sistema-nacional-de-estatisticas-de-seguranca-publica

Uso
---
python etl_risco_sinesp.py --input basededadosvde.zip --output atlas_risco_municipios.csv
python etl_risco_sinesp.py --input pasta_extraida --months 12

O script tenta reconhecer tanto layouts LONGOS
(Indicador, UF, Município, Ano, Mês, Ocorrências) quanto layouts LARGOS
(colunas separadas para Roubo de Carga, Roubo de Veículo e Furto de Veículo).
Quando não houver granularidade municipal para um registro, gera linha de
nível UF, que pode ser usada como proxy no site. O site sinaliza esse fallback.
"""
from __future__ import annotations

import argparse
import csv
import math
import re
import shutil
import tempfile
import unicodedata
import zipfile
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Iterable, Iterator

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("Dependência ausente: pip install openpyxl") from exc

TARGETS = {
    "ROUBO DE CARGA": "roubo_carga",
    "ROUBO DE CARGAS": "roubo_carga",
    "ROUBO DE VEICULO": "roubo_veiculo",
    "ROUBO DE VEICULOS": "roubo_veiculo",
    "FURTO DE VEICULO": "furto_veiculo",
    "FURTO DE VEICULOS": "furto_veiculo",
}

MONTHS = {
    "JANEIRO": 1, "JAN": 1, "FEVEREIRO": 2, "FEV": 2, "MARCO": 3, "MAR": 3,
    "ABRIL": 4, "ABR": 4, "MAIO": 5, "MAI": 5, "JUNHO": 6, "JUN": 6,
    "JULHO": 7, "JUL": 7, "AGOSTO": 8, "AGO": 8, "SETEMBRO": 9, "SET": 9,
    "OUTUBRO": 10, "OUT": 10, "NOVEMBRO": 11, "NOV": 11, "DEZEMBRO": 12, "DEZ": 12,
}

UF_RE = re.compile(r"^[A-Z]{2}$")


def norm(value) -> str:
    s = "" if value is None else str(value)
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"\s+", " ", s.strip().upper())
    return s


def parse_number(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value) if math.isfinite(float(value)) else 0.0
    s = str(value).strip().replace(" ", "")
    s = re.sub(r"[^0-9,.-]", "", s)
    if not s:
        return 0.0
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_year(value) -> int | None:
    try:
        y = int(float(str(value).strip()))
        return y if 2000 <= y <= 2100 else None
    except Exception:
        return None


def parse_month(value) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        m = int(value)
        return m if 1 <= m <= 12 else None
    s = norm(value)
    if s.isdigit():
        m = int(s)
        return m if 1 <= m <= 12 else None
    return MONTHS.get(s)


def find_col(headers: list[str], *candidates: str) -> int:
    nh = [norm(h) for h in headers]
    for candidate in candidates:
        c = norm(candidate)
        for i, h in enumerate(nh):
            if h == c:
                return i
    for candidate in candidates:
        c = norm(candidate)
        for i, h in enumerate(nh):
            if c in h:
                return i
    return -1


def detect_target(text: str) -> str | None:
    t = norm(text)
    for label, key in TARGETS.items():
        if label in t:
            return key
    return None


def discover_input(input_path: Path) -> tuple[list[Path], tempfile.TemporaryDirectory | None]:
    tmp = None
    root = input_path
    if input_path.is_file() and input_path.suffix.lower() == ".zip":
        tmp = tempfile.TemporaryDirectory(prefix="atlas_sinesp_")
        root = Path(tmp.name)
        with zipfile.ZipFile(input_path) as zf:
            zf.extractall(root)
    if input_path.is_file() and input_path.suffix.lower() in {".xlsx", ".xlsm", ".csv"}:
        return [input_path], tmp
    files = [p for p in root.rglob("*") if p.is_file() and p.suffix.lower() in {".xlsx", ".xlsm", ".csv"}]
    return sorted(files), tmp


def iter_csv_rows(path: Path) -> Iterator[tuple[str, list, Iterable[list]]]:
    # tenta UTF-8, depois Latin-1; detecta ; / , pelo cabeçalho
    for enc in ("utf-8-sig", "latin-1"):
        try:
            f = path.open("r", encoding=enc, errors="strict", newline="")
            first = f.readline()
            f.seek(0)
            delimiter = ";" if first.count(";") >= first.count(",") else ","
            reader = csv.reader(f, delimiter=delimiter)
            rows = iter(reader)
            header = next(rows, [])
            yield path.name, header, rows
            f.close()
            return
        except UnicodeDecodeError:
            continue


def iter_xlsx_rows(path: Path) -> Iterator[tuple[str, list, Iterable[list]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        # procura cabeçalho útil nas primeiras 30 linhas
        cache = []
        header_idx = None
        header = None
        rows_iter = ws.iter_rows(values_only=True)
        for idx, row in enumerate(rows_iter, start=1):
            vals = ["" if v is None else v for v in row]
            cache.append(vals)
            joined = " | ".join(norm(v) for v in vals)
            if any(k in joined for k in ("INDICADOR", "ROUBO DE CARGA", "MUNICIPIO", "OCORRENCIA")) and any(norm(v) in {"UF", "UNIDADE DA FEDERACAO"} or "UF" == norm(v) for v in vals):
                header_idx = idx
                header = vals
                break
            if idx >= 30:
                break
        if not header:
            continue
        def chained():
            for row in rows_iter:
                yield ["" if v is None else v for v in row]
        yield f"{path.name}::{ws.title}", header, chained()
    wb.close()


def rows_from_file(path: Path):
    if path.suffix.lower() == ".csv":
        yield from iter_csv_rows(path)
    else:
        yield from iter_xlsx_rows(path)


def process_table(source: str, header: list, rows: Iterable[list], records: list[dict]):
    hnorm = [norm(h) for h in header]
    ix_uf = find_col(header, "UF", "SIGLA UF", "UNIDADE DA FEDERACAO")
    ix_mun = find_col(header, "MUNICIPIO", "MUNICÍPIO", "NOME MUNICIPIO")
    ix_ind = find_col(header, "INDICADOR", "NATUREZA", "TIPO INDICADOR", "EVENTO")
    ix_year = find_col(header, "ANO")
    ix_month = find_col(header, "MES", "MÊS")
    ix_value = find_col(header, "OCORRENCIAS", "OCORRÊNCIAS", "TOTAL OCORRENCIAS", "VALOR", "QUANTIDADE", "TOTAL")

    wide = {}
    for idx, h in enumerate(hnorm):
        target = detect_target(h)
        if target:
            wide[target] = idx

    if ix_uf < 0:
        return

    for row in rows:
        if not row:
            continue
        def val(i):
            return row[i] if i >= 0 and i < len(row) else ""
        uf = norm(val(ix_uf))
        if not UF_RE.match(uf):
            continue
        mun = str(val(ix_mun)).strip() if ix_mun >= 0 else ""
        year = parse_year(val(ix_year)) if ix_year >= 0 else None
        month = parse_month(val(ix_month)) if ix_month >= 0 else None
        period = (year, month or 12) if year else None
        if wide:
            for target, ix in wide.items():
                number = parse_number(val(ix))
                records.append({"source": source, "uf": uf, "municipio": mun, "target": target, "value": number, "period": period})
            continue
        if ix_ind < 0 or ix_value < 0:
            continue
        target = detect_target(str(val(ix_ind)))
        if not target:
            continue
        number = parse_number(val(ix_value))
        records.append({"source": source, "uf": uf, "municipio": mun, "target": target, "value": number, "period": period})


def aggregate(records: list[dict], months: int):
    # Se houver datas, seleciona a janela nacional dos N meses mais recentes encontrados.
    periods = sorted({r["period"] for r in records if r["period"]})
    selected_periods = set(periods[-months:]) if periods else None
    period_start = selected_periods and min(selected_periods)
    period_end = selected_periods and max(selected_periods)

    agg = defaultdict(lambda: {"roubo_carga": 0.0, "roubo_veiculo": 0.0, "furto_veiculo": 0.0, "months": set(), "sources": set()})
    for r in records:
        if selected_periods and r["period"] and r["period"] not in selected_periods:
            continue
        level = "MUNICIPIO" if r["municipio"] else "UF"
        key = (level, r["uf"], norm(r["municipio"]) if r["municipio"] else "")
        a = agg[key]
        a[r["target"]] += max(0.0, r["value"])
        if r["period"]:
            a["months"].add(r["period"])
        a["sources"].add(r["source"])

    out = []
    for (level, uf, mun_norm), a in agg.items():
        # Prioriza linhas municipais. Linhas UF servem de fallback quando o dado municipal não existir.
        out.append({
            "nivel": level,
            "municipio": mun_norm.title() if mun_norm else "",
            "uf": uf,
            "roubo_carga": round(a["roubo_carga"]),
            "roubo_veiculo": round(a["roubo_veiculo"]),
            "furto_veiculo": round(a["furto_veiculo"]),
            "periodo_inicio": f"{period_start[0]:04d}-{period_start[1]:02d}" if period_start else "",
            "periodo_fim": f"{period_end[0]:04d}-{period_end[1]:02d}" if period_end else "",
            "meses_observados": len(a["months"]) if a["months"] else "",
            "fonte": "MJSP / Sinesp VDE",
            "qualidade": "MUNICIPAL" if level == "MUNICIPIO" else "PROXY_UF",
        })
    out.sort(key=lambda x: (x["nivel"] != "MUNICIPIO", x["uf"], x["municipio"]))
    return out


def write_output(rows: list[dict], output: Path):
    output.parent.mkdir(parents=True, exist_ok=True)
    fields = ["nivel", "municipio", "uf", "roubo_carga", "roubo_veiculo", "furto_veiculo", "periodo_inicio", "periodo_fim", "meses_observados", "fonte", "qualidade"]
    with output.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, delimiter=";")
        w.writeheader()
        w.writerows(rows)


def main():
    ap = argparse.ArgumentParser(description="Normaliza Sinesp VDE para o Atlas Market Intelligence.")
    ap.add_argument("--input", required=True, type=Path, help="ZIP/XLSX/CSV do Sinesp VDE ou pasta extraída")
    ap.add_argument("--output", default=Path("atlas_risco_municipios.csv"), type=Path)
    ap.add_argument("--months", type=int, default=12, help="janela dos meses mais recentes (padrão 12)")
    args = ap.parse_args()
    if not args.input.exists():
        raise SystemExit(f"Entrada não encontrada: {args.input}")
    files, tmp = discover_input(args.input)
    if not files:
        raise SystemExit("Nenhum XLSX/CSV encontrado na entrada.")
    records = []
    print(f"Arquivos encontrados: {len(files)}")
    try:
        for path in files:
            print(f"Lendo: {path.name}")
            for source, header, rows in rows_from_file(path):
                process_table(source, header, rows, records)
    finally:
        if tmp:
            tmp.cleanup()
    if not records:
        raise SystemExit("Nenhum registro de Roubo de Carga / Roubo de Veículo / Furto de Veículo foi reconhecido. Verifique o layout.")
    result = aggregate(records, max(1, args.months))
    write_output(result, args.output)
    municipal = sum(1 for r in result if r["nivel"] == "MUNICIPIO")
    uf = sum(1 for r in result if r["nivel"] == "UF")
    print(f"Concluído: {len(result)} linhas | municipais={municipal} | UF={uf}")
    print(f"Arquivo: {args.output}")


if __name__ == "__main__":
    main()
